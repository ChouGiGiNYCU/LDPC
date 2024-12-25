import torch
import matplotlib.pyplot as plt  # 繪圖
from UseFunction import *
from collections import defaultdict
from parity_check_matrix import *
from sparse_bp_model import SparseBPNeuralNetwork
import pandas
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np 
import pandas as pd
def write_cn2vn_weight_to_txt(file_name, data):
    """
    将数据写入 txt 文件
    参数:
    file_name (str): 保存文件的路径和名称
    data (list of str): 要写入文件的数据列表，每个元素为一行的内容
    """
    with open(file_name, 'w') as file:
        for line in data:
            file.write(line + '\n')
    print(f"Data has been written to {file_name}")
def show_plot(weights_list,bins):
    """ 把每一層的參數權重都 plot 出來 """
    # 創建一個 1x2 的子圖網格 (1 行 2 列)
    # 創建一個包含四個子圖的窗口
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2)  # 2行2列
    # 繪製第一個直方圖
    n1, bins1, patches1 = ax1.hist(weights_list[0], bins=bins, edgecolor='black', alpha=0.7)
    ax1.set_title('Weights Distribution 1')
    ax1.set_xlabel('Weight')
    ax1.set_ylabel('Frequency')

    # 繪製第二個直方圖
    n2, bins2, patches2 = ax2.hist(weights_list[1], bins=bins, edgecolor='black', alpha=0.7)
    ax2.set_title('Weights Distribution 2')
    ax2.set_xlabel('Weight')
    ax2.set_ylabel('Frequency')

    # 繪製第三個直方圖
    n3, bins3, patches3 = ax3.hist(weights_list[2], bins=bins, edgecolor='black', alpha=0.7)
    ax3.set_title('Weights Distribution 3')
    ax3.set_xlabel('Weight')
    ax3.set_ylabel('Frequency')

    # 繪製第四個直方圖
    n4, bins4, patches4 = ax4.hist(weights_list[3], bins=bins, edgecolor='black', alpha=0.7)
    ax4.set_title('Weights Distribution 4')
    ax4.set_xlabel('Weight')
    ax4.set_ylabel('Frequency')

    # 設置所有子圖的 y 軸範圍
    max_freq = max(max(n1), max(n2), max(n3), max(n4))  # 獲取所有圖表中的最大頻率
    for ax in [ax1, ax2, ax3, ax4]:
        ax.set_ylim(0, max_freq + 1)  # 增加一些空間以顯示條形邊框

    # 調整子圖間距
    plt.tight_layout()

    # 顯示圖表
    plt.show()
def read_girth_num(file_name):
    cycle_dict = {}
    with open(file_name,'r') as file:
        for idx,line in enumerate(file):
            if idx==0:
                girth = int(line.strip())
            else:
                CN = int(line.strip().split()[0])
                VN = int(line.strip().split()[1])
                number = int(line.strip().split()[2])
                cycle_dict[(CN,VN)] = number
    return girth,cycle_dict
def save_model_params_to_excel(model, file_name):
    with pd.ExcelWriter(file_name,engine='openpyxl') as writer:
        for name,param in model.named_parameters():
            param_data = pd.DataFrame(param.data.cpu().numpy())
            param_data.to_excel(writer,sheet_name=name[:30])
    print("Write Model All Weight into Excel")
def write_bias_weight_to_txt(model,file_name):
    lines = ["5 96 48"]
    line_str = ""
    for it, param in enumerate(model.bias_scaling_vectors):
        bias_matrix = param.detach().cpu().numpy()  # 轉換為 numpy array 以便讀取
        for vn in range(len(bias_matrix)):
            line_str = f"{it} {vn} {bias_matrix[vn][vn]}"  # iteration  VN   bias_weight
            lines.append(line_str)
    with open(file_name,'w') as f:
        for line in lines:
            f.write(line+'\n')
    print("Write All bias to txt file success !!")       
def write_vn2cn_mean_to_txt(file_name,data):
    with open(file_name, 'w') as file:
        for line in data:
            file.write(line + '\n')
    print(f"Data has been written to {file_name}")
    

################ color setting  ################
colors = [ # 根據每0.5設置的顏色範圍
    'FFF5E1D7',  # -1 to -0.75 淺膚色
    'FFEED6BF',  # -0.75 to -0.5
    'FFE5CDA6',  # -0.5 to -0.25
    'FFD9B492',  # -0.25 to 0
    'FFCF9F7F',  # 0 to 0.25
    'FFD19C6C',  # 0.25 to 0.5 中間膚色
    'FFC6895A',  # 0.5 to 0.75
    'FFB87B48',  # 0.75 to 1
    'FFAA6F36',  # 1 to 1.25
    'FF965E2D',  # 1.25 to 1.5
    'FF874D21',  # 1.5 to 1.75
    'FF8B4513'   # 1.75 to 2 深膚色
]
################################################

################# Some parameter setting #################
Load_model_file_name = r'BPNN_Mulitiloss_H9648_8_valid.pth'
Load_setting = True
# device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
iteration = 4 # iteration number
Excel_file_name = "H_96_48_ZeroCodeWord.xlsx" # save 所有vn update的weight each iteration 
All_weight_record_excel_file_name = "H_96_48_ModelAllWeight.xlsx" # save all model weight in excel 
cn2vn_weight_txt_file_name = "H_96_48_weight_cn2vn.txt" # save final layer weight
weight_txt_vn2cn_file_name = "H_96_48_weight_vn2cn_mean.txt" # save vn->cn 所有edge的mean值
final_layer_file_name = "H_96_48_Final_Layer_weight.txt" # save final layer weight
bias_weight_txt_file_name = "H_96_48_bias.txt"
cycle_file_name = "H_96_48_cycle_num_record.txt" # read record_cycle.txt file
H_file_name = "H_96_48.txt" # parity_check matrix
weight_data_cn2vn_str = f"{iteration} 96 48" # iteration-VN_node-CN_node  
wb = openpyxl.Workbook()
ws = wb.active
bins = np.arange(0, 1.5, 0.01)  
weight_data_cn2vn = []
weight_data_vn2cn = []
final_layer_weight = []
weight_data_cn2vn.append(weight_data_cn2vn_str) # it - cn - vn - value
weight_data_vn2cn.append(weight_data_cn2vn_str)
girth , cycle_dict = read_girth_num(cycle_file_name)
################# Some parameter setting #################

# parity_check_matrix = get_parity_check_matrix()
parity_check_matrix = read_H_file(H_file_name)
parity_check_matrix_colsize = len(parity_check_matrix[0])
parity_check_matrix_rowsize = len(parity_check_matrix)
# first hidden layer node number count 1 number in parity_matrix
hidden_layer_node = sum(sum(row) for row in parity_check_matrix)
print("parity_check_matrix_rowsize : ",parity_check_matrix_rowsize)
print("parity_check_matrix_colsize : ",parity_check_matrix_colsize)
print("hidden_layer_node : ",hidden_layer_node)

# define model paramter
input_size = parity_check_matrix_colsize
hidden_layer_node = hidden_layer_node
output_size = parity_check_matrix_colsize


##################### Construct CN2VN #####################
CN2VN_pos = []
max_row_degree =0 
for r in range(0,len(parity_check_matrix)):
    tmp=[]
    for c in range(0,len(parity_check_matrix[0])):
        if parity_check_matrix[r][c]==1:
            tmp.append(c)
    max_row_degree = max(max_row_degree,len(tmp))
    CN2VN_pos.append(tmp)
# print(CN2VN_pos)
# CN2VN_pos
print("max_row_degree : ",max_row_degree)
###########################################################

##################### Construct VN2CN #####################
VN2CN=[]
max_col_degree=0
for c in range(0,len(parity_check_matrix[0])):
    tmp =[]
    for r in range(0,len(parity_check_matrix)):
        if parity_check_matrix[r][c]==1:
            tmp.append(r)
    VN2CN.append(tmp)
    max_col_degree = max(max_col_degree,len(tmp))
# print(VN2CN)
print("max_col_degree : ",max_col_degree)
###########################################################

##################### Write parity_check_matrix to txt  #####################
# Write_txt_File(file_name, parity_check_matrix_colsize, parity_check_matrix_rowsize, max_col_degree, max_row_degree, VN2CN, CN2VN_pos)
##################### Write parity_check_matrix to txt  #####################

############################################################ Construct Every Layer Mask ############################################################
############################ First layer mask (Input to Cn Update - Vns_to_Cn ) ############################
connect_first_layer = []
dict_update_cn = {}
dict_update_cn_arr = defaultdict(list)
dict_vn_pos = defaultdict(list)
for CN in range(0,len(CN2VN_pos)):
    for c in range(0,len(CN2VN_pos[CN])):
        Vns_to_Cn = CN2VN_pos[CN][0:c] + CN2VN_pos[CN][c+1:len(CN2VN_pos[CN])]
        connect_first_layer.append(Vns_to_Cn)
        dict_update_cn[tuple(Vns_to_Cn)]=(CN2VN_pos[CN][c],CN) # [VNs connect to CN ] - update(CN->CN2VN_pos[r][c]) (VN,CN)
        dict_update_cn_arr[(CN2VN_pos[CN][c],CN)].append(tuple(Vns_to_Cn))
        # print(dict_update_cn[tuple(Vns_to_Cn)], "-",tuple(Vns_to_Cn))

# construct first layer mask (CN update)
first_layer_mask = [[False]*parity_check_matrix_colsize for i in range(0,hidden_layer_node)]
print("layer_mask rowsize is :",len(first_layer_mask)," , colsize : ",len(first_layer_mask[0]))
for i in range(0,len(connect_first_layer)):
    for num in connect_first_layer[i]:
        first_layer_mask[i][num]=True
############################ First layer mask (Input to Cn Update - Vns_to_Cn ) ############################

############################ Second layer mask (CNs to VNi) ############################
# save neural network node position ([connect to VNs]- node_position)
dict_connect_first_layer ={}
for idx,arr in enumerate(connect_first_layer):
    dict_connect_first_layer[tuple(arr)]=idx 
# 找出input到firsy_layer連接中，firsy_layer(VN update)的node哪些是CNx update此VN
for key in sorted(dict_update_cn_arr.keys(),key=lambda x:x[1]):
    VN = key[0]
    for arr in dict_update_cn_arr[key]:
        dict_vn_pos[VN].append(dict_connect_first_layer[arr])
# sort 
sorted_keys = sorted(dict_vn_pos.keys(),key=lambda x:x)
dict_vn_pos_sort = {key: dict_vn_pos[key] for key in sorted_keys}
dict_vn_pos=dict_vn_pos_sort.copy()
# show update key with node
# for key in dict_vn_pos.keys():
    # print(key,"-",dict_vn_pos[key])
# construct CN2VN mask - second layer connect table matrix(mask)
CN2VN_mask_VNUpdate = [[False]*hidden_layer_node for i in range(0,hidden_layer_node)]
idx=0
tmp=[]
for VN in dict_vn_pos.keys():
    if(len(dict_vn_pos[VN])==1):
        idx+=1
        tmp.append([VN2CN[VN][0],VN])
        continue
    for i,node1 in enumerate(dict_vn_pos[VN]):
        for node2 in dict_vn_pos[VN]:
            if node1!=node2:
                CN2VN_mask_VNUpdate[idx][node2]=True
        idx+=1
        tmp.append([VN2CN[VN][i],VN]) # CN,VN
true_indices_per_row = [[idx for idx, val in enumerate(row) if val] for row in CN2VN_mask_VNUpdate]
VN2CN_dict = {}
for idx,arr in enumerate(true_indices_per_row):
    print("neourn : {} - {} | VN {} -> CN {}".format(idx, arr, tmp[idx][1], tmp[idx][0]))
    neourn = idx
    tmp_dict={
        "VN":tmp[idx][1],
        "CN":tmp[idx][0],
        "connect_neourn":arr
    }
    VN2CN_dict[neourn] = tmp_dict
############################ Second layer mask (CNs to VNi) ############################

############################ Third layer mask (VNs to CNi) ############################
#constuct third mask of VNs->CNi update (record VN->CN node postion in neural network)
dict_vn_to_cn_node = defaultdict(list)
idx=0
for VN in dict_vn_pos.keys():
    for CN in VN2CN[VN]:
        dict_vn_to_cn_node[(VN,CN)] = idx
        idx+=1
# print("(VN update CN) - Network node idx")
# for key in dict_vn_to_cn_node.keys():
#     print(key,"-",dict_vn_to_cn_node[key])
#construct third mask(VN->CN)
VN2CN_mask_CNUpdate = [[False]*hidden_layer_node for i in range(0,hidden_layer_node)]
# 照著cn0->vn1 -> cn0->vn2 ...順序
idx=0
CN_2_VN_record=[]
for CN in range(0,parity_check_matrix_rowsize):
    #  CN->VN1 update
    for VN1 in  CN2VN_pos[CN]: 
        for VN2 in CN2VN_pos[CN]:
            if VN1!=VN2:
                # print(dict_vn_to_cn_node[(VN2,CN)])
                VN2CN_mask_CNUpdate[idx][dict_vn_to_cn_node[(VN2,CN)]]=True
        idx+=1
        CN_2_VN_record.append([CN,VN1])
true_indices_per_row = [[idx for idx, val in enumerate(row) if val] for row in VN2CN_mask_CNUpdate]
CN2VN_dict={}
for idx,arr in enumerate(true_indices_per_row):
    # print("node : ",idx," - ",arr,"| CN",CN_2_VN_record[idx][0],"->","VN",CN_2_VN_record[idx][1])
    print("neourn : {} - {} | CN {} -> VN {}".format(idx, arr, CN_2_VN_record[idx][0], CN_2_VN_record[idx][1]))
    neourn = idx 
    tmp_dict = {
        "VN" : CN_2_VN_record[idx][1],
        "CN" : CN_2_VN_record[idx][0],
        "connect_neourn" : arr
    }
    CN2VN_dict[neourn]=tmp_dict
############################ Third layer mask (VNs to CNi) ############################

############################ Final layer mask (CNs to output) ############################
# construct fourth mask of  CN ->VN + channel LLR = output
dict_cn_to_output = defaultdict(list)
idx=0
for CN in range(0,len(CN2VN_pos)):
    for VN in CN2VN_pos[CN]:
        dict_cn_to_output[VN].append(idx)
        idx+=1
# sort 
sorted_keys = sorted(dict_cn_to_output.keys(),key=lambda x:x)
dict_cn_to_output_sort = {key: dict_cn_to_output[key] for key in sorted_keys}
dict_cn_to_output=dict_cn_to_output_sort.copy()
# for key in dict_cn_to_output.keys():
#     print(key,"-",dict_cn_to_output[key]) # 跟第一層一樣
# construct fouth final CN update to result
CN2VN_mask_output = [[False]*hidden_layer_node for i in range(0,parity_check_matrix_colsize)]
for VN in range(0,len(dict_cn_to_output)):
    for node in dict_cn_to_output[VN]:
        CN2VN_mask_output[VN][node] = True
    # print(CN2VN_mask_output[VN])
    
true_indices_per_row = [[idx for idx, val in enumerate(row) if val] for row in CN2VN_mask_output]
# for idx,arr in enumerate(true_indices_per_row):
#     print("neourn : {} - {} ".format(idx, arr))
############################ Final layer mask (CNs to output) ############################

############################################################ Construct Every Layer Mask ############################################################

###########################  Dot_Bias_Matrix ###########################
# 神經網路裡面的bias=[LLR1 ... LLR_VN]*matrix
dot_bias_matrix=torch.zeros(parity_check_matrix_colsize ,hidden_layer_node)
idx=0
for i in range(0,len(VN2CN)):
    for j in range(0,len(VN2CN[i])):
        dot_bias_matrix[i][idx]=1
        idx = idx + 1
####################### setting model ###################################
model = SparseBPNeuralNetwork(input_size,hidden_layer_node,output_size,first_layer_mask,CN2VN_mask_VNUpdate,VN2CN_mask_CNUpdate,CN2VN_mask_output,dot_bias_matrix)
if Load_setting:
    state_dict = torch.load(Load_model_file_name,map_location=torch.device('cpu'))  
    model.load_state_dict(state_dict)
save_model_params_to_excel(model,All_weight_record_excel_file_name)  
write_bias_weight_to_txt(model,bias_weight_txt_file_name)
########################################################################

###########################  Write Txt file for hidden layer of VN update  ###########################
# 使用 named_modules() 来获取模型中的所有模块
fc_layers = []
for name, module in model.named_modules():
    if name in ['fc1', 'fc3', 'fc5', 'fc7']:
        fc_layers.append((name, module))
# 打印所有 fc(VN update Layer) 层的权重和偏置
layer_data = []
for name, fc_layer in fc_layers:
    weight_list = fc_layer.weight.data.tolist()  # 将 Tensor 转换为普通列表
    print(f"weight_list size : {len(weight_list)} x {len(weight_list[0])}")    
    layer_data.append(weight_list)

weights_list  = [[] for _ in range(iteration)]
#  write the  it
for it in range(iteration):
    ws.cell(row=1, column=it+3, value=f"Iteration-{it}")
curr_row=2
VN2CN_col=1
CN2VN_col=2
weight_data_lines = ["4 96 48"]
weight_data_vn2cn_str = ""
for key,dict in VN2CN_dict.items():
    currVN = dict["VN"]
    currCN = dict["CN"]
    node =   dict["connect_neourn"] # arr
    ws.cell(row=curr_row,column=VN2CN_col,value=f"VN{currVN} -> CN{currCN}")
    # print(f"VN{currVN} -> CN{currCN}")
    curr_row+=1
    if len(node)==0:
        pass
    else:
        mean=[[] for _ in range(iteration)]
        for neourn in node:
            prev_cn = CN2VN_dict[neourn]["CN"]
            prev_vn = CN2VN_dict[neourn]["VN"]
            ws.cell(row=curr_row,column=CN2VN_col,value=f"CN{prev_cn} -> VN{prev_vn}")
            # print(f"CN{prev_cn} -> VN{prev_vn}")
            # write data
            for it in  range(iteration):
                print(f"{prev_cn}->{prev_vn}->{currVN}->{currCN}")
                weight_data_cn2vn_str = f"{it} {prev_cn} {prev_vn} {currCN} {layer_data[it][key][neourn]}"
                weight_data_cn2vn.append(weight_data_cn2vn_str)
                mean[it].append(layer_data[it][key][neourn])
                cell=ws.cell(row=curr_row,column=CN2VN_col+it+1,value=f"{layer_data[it][key][neourn]}")
                ######################
                value = layer_data[it][key][neourn]
                # 計算顏色索引
                index = int((value + 1) // 0.5)  # 確保索引在正確範圍內
                index = max(0, min(index, len(colors) - 1))  # 防止超出顏色範圍
                fill_color = colors[index]
                # 設定背景顏色
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.fill = fill
                ######################
                # print(f"it - {it} data : {layer_data[it][key][neourn]}")
                weights_list[it].append(layer_data[it][key][neourn])
            if (prev_cn,prev_vn) in cycle_dict:
                ws.cell(row=curr_row,column=CN2VN_col+iteration+1,value=f"cycle:{girth} , number :{cycle_dict[(prev_cn,prev_vn)]}")
            curr_row+=1
        for it in range(iteration):
            weight_data_vn2cn_str = f"{it} {currVN} {currCN} {sum(mean[it])/len(mean[it])}"
            weight_data_lines.append(weight_data_vn2cn_str)
wb.save(Excel_file_name)

write_cn2vn_weight_to_txt(cn2vn_weight_txt_file_name,weight_data_cn2vn)
write_vn2cn_mean_to_txt(weight_txt_vn2cn_file_name,weight_data_lines)
show_plot(weights_list,bins)
#######################################

##################################### Write Final Layer weight Txt file #####################################   

fc_layers = []
for name, module in model.named_modules():
    if name in ['fc9']:
        fc_layers.append((name, module))
layer_data = []
for name, fc_layer in fc_layers:
    weight_list = fc_layer.weight.data.tolist()  # 将 Tensor 转换为普通列表
    print(f"weight_list size : {len(weight_list)} x {len(weight_list[0])}")   
    layer_data.append(weight_list)
final_layer_weight_str = ""
final_layer_weight.append(f"{parity_check_matrix_colsize} {parity_check_matrix_rowsize}") # cnx -> vnx -> weight
for neourn,dict in CN2VN_dict.items():
    curr_cn =  dict["CN"]
    curr_vn =  dict["VN"]
    # print(f"VN : {curr_vn} | CN : {curr_cn} | neourn : {neourn}")
    weight  =  layer_data[0][curr_vn][neourn]
    if weight ==0:
        print("error !!!!")
        exit()
    final_layer_weight_str = f"{curr_cn} {curr_vn} {weight}"
    final_layer_weight.append(final_layer_weight_str)

with open(final_layer_file_name,'w') as file:
    for final_layer_weight_str in final_layer_weight:
        file.write(final_layer_weight_str)
        file.write("\n") 
print("Write Final Layer data success")
##############################################################################################################
